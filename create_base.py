from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment


class CreateTable:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def add_to(self):
        pass

    def cell_format(self, place):
        self.ws[place].alignment = Alignment(horizontal='center', vertical='center')
        t_font = Font(name='Calibri', size=16, italic=True, color=colors.BLACK, bold=True)
        # 给font属性赋值font对象即可
        self.ws[place].font = t_font


if __name__ == '__main__':
    cr = CreateTable()
