from openpyxl import load_workbook


class CheckList:

    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def row_col(self):
        """
        字典返回表格数据
        :return:
        """
        key = []
        for k in self.ws[1]:
            key.append(k.value)
        for r in range(2, self.ws.max_row+1):
            va = {}
            for c in range(1, self.ws.max_column+1):
                cell_v = self.ws.cell(r, c).value
                va[key[c-1]] = cell_v
            yield va


if __name__ == '__main__':
    ch = CheckList('仓库.xlsx')
    for i in ch.row_col():
        print(i)

